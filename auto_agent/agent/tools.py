"""
OmniAgent X - ULTIMATE Tools Engine
====================================
ALL capabilities the agent can use - NO LIMITATIONS
"""
import os
import subprocess
import json
import time
import logging
import shutil
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
import requests
from bs4 import BeautifulSoup
from datetime import datetime
from config import settings

logger = logging.getLogger(__name__)

# Try to import optional libraries
try:
    import pyautogui
    PYAUTOGUI_AVAILABLE = True
except:
    PYAUTOGUI_AVAILABLE = False
    logger.warning("PyAutoGUI not available - mouse/keyboard control limited")

try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.chrome.options import Options
    SELENIUM_AVAILABLE = True
except:
    SELENIUM_AVAILABLE = False
    logger.warning("Selenium not available - browser automation limited")


class ToolsEngine:
    """
    ULTIMATE All tools and capabilities - NO LIMITATIONS
    """
    
    def __init__(self):
        self.results_cache = {}
        self.browser_driver = None
        logger.info("🔧 ULTIMATE Tools Engine initialized - NO LIMITS")
    
    # ==================== FILE OPERATIONS - NO LIMITS ====================
    
    def read_file(self, filepath: str) -> str:
        """Read a file and return its contents - NO LIMITS"""
        try:
            path = Path(filepath).absolute()
            if not path.exists():
                return f"❌ Fayl topilmadi: {filepath}"
            
            with open(path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Show full content for large files too
            logger.info(f"📄 Read file: {filepath} ({len(content)} chars)")
            return f"✅ Fayl o'qildi: {filepath}\n\nHajmi: {len(content)} belgi\n\n--- MAZMUNI ---\n{content}"
        
        except Exception as e:
            logger.error(f"❌ Error reading file: {e}")
            return f"❌ Xatolik: {str(e)}"
    
    def read_file_binary(self, filepath: str) -> str:
        """Read binary file info"""
        try:
            path = Path(filepath)
            if not path.exists():
                return f"❌ Fayl topilmadi: {filepath}"
            
            stat = path.stat()
            return f"📄 Fayl: {filepath}\n- Hajm: {stat.st_size} bytes\n- Yaratilgan: {datetime.fromtimestamp(stat.st_ctime)}\n- O'zgartirilgan: {datetime.fromtimestamp(stat.st_mtime)}"
        
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def write_file(self, filepath: str, content: str) -> str:
        """Write content to a file"""
        try:
            path = Path(filepath)
            path.parent.mkdir(parents=True, exist_ok=True)
            
            with open(path, 'w', encoding='utf-8') as f:
                f.write(content)
            
            logger.info(f"✍️ Wrote file: {filepath}")
            return f"✅ Fayl yaratildi: {filepath}\n\nHajmi: {len(content)} belgi"
        
        except Exception as e:
            logger.error(f"❌ Error writing file: {e}")
            return f"❌ Xatolik: {str(e)}"
    
    def delete_file(self, filepath: str) -> str:
        """Delete a file"""
        try:
            path = Path(filepath)
            if not path.exists():
                return f"❌ Fayl topilmadi: {filepath}"
            
            path.unlink()
            logger.info(f"🗑️ Deleted file: {filepath}")
            return f"✅ Fayl o'chirildi: {filepath}"
        
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def list_directory(self, path: str = ".") -> str:
        """List files in a directory"""
        try:
            target = Path(path)
            if not target.exists():
                return f"❌ Papka topilmadi: {path}"
            
            items = []
            for item in target.iterdir():
                item_type = "📁" if item.is_dir() else "📄"
                size = f" ({item.stat().st_size} bytes)" if item.is_file() else ""
                items.append(f"{item_type} {item.name}{size}")
            
            result = f"📂 Papka: {path}\n\n" + "\n".join(items)
            logger.info(f"📂 Listed directory: {path}")
            return result
        
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def search_files(self, directory: str, pattern: str) -> str:
        """Search for files matching a pattern"""
        try:
            target = Path(directory)
            if not target.exists():
                return f"❌ Papka topilmadi: {directory}"
            
            matches = []
            for item in target.rglob(pattern):
                matches.append(f"📄 {item.relative_to(target)}")
            
            if not matches:
                return f"🔍 '{pattern}' bo'yicha hech narsa topilmadi"
            
            result = f"🔍 Natijalar ({len(matches)} ta):\n" + "\n".join(matches[:20])
            if len(matches) > 20:
                result += f"\n... va yana {len(matches)-20} ta"
            
            return result
        
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    # ==================== TERMINAL COMMANDS ====================
    
    def execute_command(self, command: str) -> str:
        """Execute a terminal command"""
        try:
            logger.info(f"⚡ Executing command: {command}")
            
            # Security: block dangerous commands
            dangerous = ["rm -rf /", "format", "del /", "dd if="]
            if any(cmd in command.lower() for cmd in dangerous):
                return "❌ Xavfli buyruq bloklandi"
            
            result = subprocess.run(
                command,
                shell=True,
                capture_output=True,
                text=True,
                timeout=settings.MAX_CODE_EXECUTION_TIME
            )
            
            output = result.stdout if result.stdout else result.stderr
            
            if not output:
                output = "✅ Buyruq muvaffaqiyatli bajarildi"
            
            # Limit output
            if len(output) > 3000:
                output = output[:3000] + "\n\n... (chiqish qisqartirildi)"
            
            logger.info(f"✅ Command executed successfully")
            return f"⚡ Natija:\n\n{output}"
        
        except subprocess.TimeoutExpired:
            return "❌ Buyruq vaqt tugadi"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    # ==================== WEB OPERATIONS ====================
    
    def web_search(self, query: str, num_results: int = 5) -> str:
        """Search the web for information"""
        try:
            logger.info(f"🌐 Searching web: {query}")
            
            # Using DuckDuckGo (no API key needed)
            url = "https://html.duckduckgo.com/html/"
            data = {"q": query}
            
            response = requests.post(url, data=data, timeout=10)
            soup = BeautifulSoup(response.text, 'html.parser')
            
            results = []
            for result in soup.select('.result')[:num_results]:
                title = result.select_one('.result__title')
                link = result.select_one('.result__url')
                snippet = result.select_one('.result__snippet')
                
                if title and link:
                    results.append(f"📌 {title.get_text(strip=True)}\n   {link.get_text(strip=True)}\n   {snippet.get_text(strip=True) if snippet else ''}")
            
            if not results:
                return "🔍 Natijalar topilmadi"
            
            result_text = f"🔍 Qidiruv natijalari: '{query}'\n\n" + "\n\n".join(results)
            logger.info(f"✅ Found {len(results)} results")
            return result_text
        
        except Exception as e:
            logger.error(f"❌ Web search error: {e}")
            return f"❌ Qidiruv xatosi: {str(e)}"
    
    def scrape_website(self, url: str, query: Optional[str] = None) -> str:
        """Extract content from a website"""
        try:
            logger.info(f"🕷️ Scraping: {url}")
            
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            }
            
            response = requests.get(url, headers=headers, timeout=15)
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Remove script and style elements
            for script in soup(["script", "style"]):
                script.decompose()
            
            # Get text content
            text = soup.get_text(separator='\n', strip=True)
            
            # If specific query, look for that content
            if query:
                paragraphs = soup.find_all('p')
                relevant = [p.get_text() for p in paragraphs if query.lower() in p.get_text().lower()]
                if relevant:
                    text = "\n".join(relevant[:5])
            
            # Limit output
            if len(text) > 4000:
                text = text[:4000] + "\n\n... (content truncated)"
            
            logger.info(f"✅ Scraped {url}")
            return f"✅ {url} dan olingan ma'lumot:\n\n{text}"
        
        except Exception as e:
            return f"❌ Scraping xatosi: {str(e)}"
    
    # ==================== CODE EXECUTION ====================
    
    def execute_code(self, code: str, language: str = "python") -> str:
        """Execute code in a sandboxed environment"""
        try:
            logger.info(f"💻 Executing {language} code")
            
            # Create a temporary file
            temp_file = Path(settings.DATA_DIR) / f"temp_code.{'py' if language == 'python' else 'js'}"
            temp_file.write_text(code, encoding='utf-8')
            
            # Execute based on language
            if language == "python":
                result = subprocess.run(
                    ["python3", str(temp_file)],
                    capture_output=True,
                    text=True,
                    timeout=settings.MAX_CODE_EXECUTION_TIME
                )
            else:
                result = subprocess.run(
                    ["node", str(temp_file)],
                    capture_output=True,
                    text=True,
                    timeout=settings.MAX_CODE_EXECUTION_TIME
                )
            
            output = result.stdout if result.stdout else result.stderr
            
            if not output:
                output = "✅ Kod muvaffaqiyatli bajarildi (natija yo'q)"
            
            # Clean up
            temp_file.unlink(missing_ok=True)
            
            logger.info(f"✅ Code executed")
            return f"💻 Kod natijasi:\n\n{output}"
        
        except subprocess.TimeoutExpired:
            return "❌ Kod vaqt tugadi"
        except Exception as e:
            return f"❌ Kod xatosi: {str(e)}"
    
    # ==================== DATA ANALYSIS ====================
    
    def analyze_data(self, filepath: str) -> str:
        """Analyze data file (JSON, CSV, etc.)"""
        try:
            path = Path(filepath)
            ext = path.suffix.lower()
            
            if ext == ".json":
                with open(path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                if isinstance(data, dict):
                    keys = list(data.keys())
                    return f"📊 JSON tahlili:\n- Keys: {keys}\n- Hajmi: {len(data)} element"
                elif isinstance(data, list):
                    return f"📊 JSON tahlili:\n- Elementlar soni: {len(data)}\n- Birinchi element: {data[0] if data else 'yoq'}"
            
            elif ext == ".csv":
                import csv
                with open(path, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                
                return f"📊 CSV tahlili:\n- Qatorlar soni: {len(rows)}\n- Ustunlar: {len(rows[0]) if rows else 0}\n- Birinchi qator: {rows[0] if rows else 'yoq'}"
            
            elif ext in [".txt", ".md"]:
                with open(path, 'r', encoding='utf-8') as f:
                    content = f.read()
                
                words = content.split()
                lines = content.split('\n')
                return f"📊 Matn tahlili:\n- So'zlar: {len(words)}\n- Qatorlar: {len(lines)}\n- Belgilar: {len(content)}"
            
            return "📊 Bu fayl turi uchun tahlil qo'llab bo'lmaydi"
        
        except Exception as e:
            return f"❌ Tahlil xatosi: {str(e)}"
    
    # ==================== APPLICATION CONTROL ====================
    
    def launch_application(self, app_name: str) -> str:
        """Launch an application"""
        try:
            if os.name == "nt":  # Windows
                subprocess.Popen(app_name)
            else:  # Mac/Linux
                subprocess.Popen(["open", "-a", app_name] if os.name == "posix" else [app_name])
            
            logger.info(f"🚀 Launched: {app_name}")
            return f"🚀 Ilova ishga tushirildi: {app_name}"
        
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    # ==================== SYSTEM INFO ====================
    
    def get_system_info(self) -> str:
        """Get system information"""
        try:
            import platform
            
            info = f"""💻 Tizim ma'lumotlari:

🖥️ Operatsion tizim: {platform.system()} {platform.release()}
🐍 Python versiyasi: {platform.python_version()}
📁 Joriy papka: {os.getcwd()}
💾 Protsessor: {platform.processor()}
🏠 Kompyuter nomi: {platform.node()}
"""
            return info
        
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def get_current_time(self) -> str:
        """Get current time and date"""
        from datetime import datetime
        now = datetime.now()
        return f"⏰ Hozirgi vaqt: {now.strftime('%H:%M:%S')}\n📅 Sana: {now.strftime('%Y-%m-%d')}"
    
    # ==================== SECURITY TOOLS ====================
    
    def check_password_strength(self, password: str) -> str:
        """Analyze password strength"""
        score = 0
        feedback = []
        
        if len(password) >= 8:
            score += 1
        else:
            feedback.append("❌ Kamida 8 ta belgi kerak")
        
        if any(c.isupper() for c in password):
            score += 1
        else:
            feedback.append("⚠️ Katta harflar qo'shing")
        
        if any(c.islower() for c in password):
            score += 1
        else:
            feedback.append("⚠️ Kichik harflar qo'shing")
        
        if any(c.isdigit() for c in password):
            score += 1
        else:
            feedback.append("⚠️ Raqamlar qo'shing")
        
        special = "!@#$%^&*()_+-=[]{}|;':\",./<>?"
        if any(c in special for c in password):
            score += 1
        else:
            feedback.append("⚠️ Maxsus belgilar qo'shing")
        
        strength = ["Juda zaif", "Zaif", "O'rta", "Kuchli", "Juda kuchli"][min(score, 4)]
        
        return f"🔐 Parol tahlili:\n- Kuch: {strength} ({score}/5)\n" + "\n".join(feedback)
    
    def ping_host(self, host: str) -> str:
        """Ping a host to check connectivity"""
        try:
            param = "-n" if os.name == "nt" else "-c"
            command = ["ping", param, "4", host]
            result = subprocess.run(command, capture_output=True, text=True)
            
            if result.returncode == 0:
                return f"✅ {host} ga ulanish muvaffaqiyatli\n\n{result.stdout[:500]}"
            else:
                return f"❌ {host} ga ulanish muvaffaqiyatsiz"
        
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    # ==================== MOUSE & KEYBOARD (FULL CONTROL) ====================
    
    def mouse_move(self, x: int, y: int) -> str:
        """Move mouse to position"""
        if not PYAUTOGUI_AVAILABLE:
            return "❌ PyAutoGUI o'rnatilmagan"
        try:
            pyautogui.moveTo(x, y)
            return f"✅ Mouse {x}, {y} ga ko'chirildi"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def mouse_click(self, x: int = None, y: int = None, button: str = "left") -> str:
        """Click mouse at position"""
        if not PYAUTOGUI_AVAILABLE:
            return "❌ PyAutoGUI o'rnatilmagan"
        try:
            if x and y:
                pyautogui.click(x, y, button=button)
            else:
                pyautogui.click(button=button)
            return f"✅ Mouse click: {button}"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def mouse_drag(self, start_x: int, start_y: int, end_x: int, end_y: int) -> str:
        """Drag mouse from start to end"""
        if not PYAUTOGUI_AVAILABLE:
            return "❌ PyAutoGUI o'rnatilmagan"
        try:
            pyautogui.moveTo(start_x, start_y)
            pyautogui.dragTo(end_x, end_y, duration=1)
            return f"✅ Mouse drag: ({start_x},{start_y}) → ({end_x},{end_y})"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def keyboard_type(self, text: str) -> str:
        """Type text with keyboard"""
        if not PYAUTOGUI_AVAILABLE:
            return "❌ PyAutoGUI o'rnatilmagan"
        try:
            pyautogui.write(text, interval=0.05)
            return f"✅ Yozildi: {text}"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def keyboard_press(self, key: str) -> str:
        """Press a key"""
        if not PYAUTOGUI_AVAILABLE:
            return "❌ PyAutoGUI o'rnatilmagan"
        try:
            pyautogui.press(key)
            return f"✅ Klavisha bosildi: {key}"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def keyboard_hotkey(self, *keys) -> str:
        """Press hotkey combination (e.g., 'ctrl', 'c')"""
        if not PYAUTOGUI_AVAILABLE:
            return "❌ PyAutoGUI o'rnatilmagan"
        try:
            pyautogui.hotkey(*keys)
            return f"✅ Hotkey: {'+'.join(keys)}"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def get_screen_size(self) -> str:
        """Get screen resolution"""
        if not PYAUTOGUI_AVAILABLE:
            return "❌ PyAutoGUI o'rnatilmagan"
        try:
            size = pyautogui.size()
            return f"📺 Ekran o'lchami: {size.width} x {size.height}"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def get_mouse_position(self) -> str:
        """Get current mouse position"""
        if not PYAUTOGUI_AVAILABLE:
            return "❌ PyAutoGUI o'rnatilmagan"
        try:
            pos = pyautogui.position()
            return f"📍 Mouse pozitsiyasi: {pos.x}, {pos.y}"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    # ==================== SCREENSHOT ====================
    
    def take_screenshot(self, filepath: str = None) -> str:
        """Take a screenshot"""
        if not PYAUTOGUI_AVAILABLE:
            return "❌ PyAutoGUI o'rnatilmagan"
        try:
            if filepath is None:
                filepath = str(settings.DATA_DIR / f"screenshot_{int(time.time())}.png")
            pyautogui.screenshot(filepath)
            return f"📸 Screenshot saqlandi: {filepath}"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def screenshot_region(self, x: int, y: int, width: int, height: int, filepath: str = None) -> str:
        """Take screenshot of specific region"""
        if not PYAUTOGUI_AVAILABLE:
            return "❌ PyAutoGUI o'rnatilmagan"
        try:
            if filepath is None:
                filepath = str(settings.DATA_DIR / f"screenshot_{int(time.time())}.png")
            img = pyautogui.screenshot(region=(x, y, width, height))
            img.save(filepath)
            return f"📸 Region screenshot saqlandi: {filepath}"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    # ==================== BROWSER AUTOMATION ====================
    
    def open_browser(self, url: str = "https://google.com") -> str:
        """Open browser with URL"""
        try:
            if not SELENIUM_AVAILABLE:
                # Fallback to system browser
                subprocess.Popen(["xdg-open", url] if os.name != "nt" else ["start", "", url])
                return f"🌐 Brauzer ochildi: {url}"
            
            # Use Selenium
            options = Options()
            options.add_argument("--headless")
            self.browser_driver = webdriver.Chrome(options=options)
            self.browser_driver.get(url)
            return f"🌐 Brauzer ochildi: {url}"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def browser_click(self, selector: str) -> str:
        """Click element in browser"""
        if not SELENIUM_AVAILABLE or not self.browser_driver:
            return "❌ Brauzer ochilmagan"
        try:
            element = self.browser_driver.find_element(By.CSS_SELECTOR, selector)
            element.click()
            return f"✅ Element bosildi: {selector}"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def browser_type(self, selector: str, text: str) -> str:
        """Type text into browser element"""
        if not SELENIUM_AVAILABLE or not self.browser_driver:
            return "❌ Brauzer ochilmagan"
        try:
            element = self.browser_driver.find_element(By.CSS_SELECTOR, selector)
            element.send_keys(text)
            return f"✅ Matn kiritildi: {text}"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def browser_get_html(self) -> str:
        """Get current page HTML"""
        if not SELENIUM_AVAILABLE or not self.browser_driver:
            return "❌ Brauzer ochilmagan"
        try:
            html = self.browser_driver.page_source[:2000]
            return f"📄 Sahifa HTML:\n{html}"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def browser_close(self) -> str:
        """Close browser"""
        try:
            if self.browser_driver:
                self.browser_driver.quit()
                self.browser_driver = None
            return "✅ Brauzer yopildi"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    # ==================== EXCEL & DATA ====================
    
    def read_excel(self, filepath: str, sheet: str = None) -> str:
        """Read Excel file"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath)
            if sheet:
                ws = wb[sheet]
            else:
                ws = wb.active
            
            # Get data
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))
            
            return f"📊 Excel fayl: {filepath}\n- Sheet: {ws.title}\n- Qatorlar: {len(data)}\n\n{str(data[:10])[:500]}..."
        except ImportError:
            return "❌ openpyxl kutubxonasi kerak: pip install openpyxl"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def write_excel(self, filepath: str, data: List[List], sheet: str = "Sheet1") -> str:
        """Write to Excel file"""
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet
            
            for row in data:
                ws.append(row)
            
            wb.save(filepath)
            return f"✅ Excel yozildi: {filepath}"
        except ImportError:
            return "❌ openpyxl kutubxonasi kerak: pip install openpyxl"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    # ==================== DATABASE ====================
    
    def query_sqlite(self, db_path: str, query: str) -> str:
        """Execute SQLite query"""
        try:
            import sqlite3
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            cursor.execute(query)
            
            if query.strip().upper().startswith("SELECT"):
                results = cursor.fetchall()
                conn.close()
                return f"📊 Natijalar ({len(results)} ta):\n{str(results)[:1000]}"
            else:
                conn.commit()
                conn.close()
                return f"✅ Query bajarildi"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    # ==================== PROCESS MANAGEMENT ====================
    
    def list_processes(self) -> str:
        """List running processes"""
        try:
            if os.name == "nt":
                result = subprocess.run(["tasklist"], capture_output=True, text=True, timeout=10)
                return f"📋 Jarayonlar:\n{result.stdout[:1500]}"
            else:
                result = subprocess.run(["ps", "aux"], capture_output=True, text=True, timeout=10)
                return f"📋 Jarayonlar:\n{result.stdout[:1500]}"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def kill_process(self, pid: int) -> str:
        """Kill a process by PID"""
        try:
            if os.name == "nt":
                subprocess.run(["taskkill", "/F", "/PID", str(pid)], timeout=10)
            else:
                subprocess.run(["kill", "-9", str(pid)], timeout=10)
            return f"✅ Jarayon o'chirildi: {pid}"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def get_process_info(self, name: str) -> str:
        """Get info about a process"""
        try:
            if os.name == "nt":
                result = subprocess.run(["tasklist", "/FI", f"IMAGENAME eq {name}"], 
                                      capture_output=True, text=True, timeout=10)
                return f"📋 {name} haqida:\n{result.stdout}"
            else:
                result = subprocess.run(["pgrep", "-f", name], capture_output=True, text=True, timeout=10)
                return f"📋 {name} PID: {result.stdout}"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    # ==================== NETWORK ====================
    
    def get_ip_info(self) -> str:
        """Get public IP information"""
        try:
            response = requests.get("https://ipinfo.io/json", timeout=10)
            data = response.json()
            return f"🌐 IP Ma'lumot:\n- IP: {data.get('ip', 'N/A')}\n- Shahar: {data.get('city', 'N/A')}\n- Mamlakat: {data.get('country', 'N/A')}\n- Provider: {data.get('org', 'N/A')}"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def port_scan(self, host: str, ports: List[int] = None) -> str:
        """Scan ports on a host"""
        if ports is None:
            ports = [80, 443, 22, 21, 25, 3306, 8080, 3000]
        
        try:
            import socket
            open_ports = []
            for port in ports:
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.settimeout(1)
                result = sock.connect_ex((host, port))
                if result == 0:
                    open_ports.append(port)
                sock.close()
            
            if open_ports:
                return f"🔓 Ochik portlar ({host}): {open_ports}"
            else:
                return f"❌ Ochik portlar topilmadi"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def check_url_status(self, url: str) -> str:
        """Check if URL is accessible"""
        try:
            response = requests.get(url, timeout=10)
            return f"✅ {url} - Status: {response.status_code}"
        except Exception as e:
            return f"❌ {url} - Xatolik: {str(e)}"
    
    # ==================== DOCKER ====================
    
    def docker_list_containers(self) -> str:
        """List Docker containers"""
        try:
            result = subprocess.run(["docker", "ps", "-a"], capture_output=True, text=True, timeout=10)
            return f"🐳 Docker konteynerlar:\n{result.stdout[:1000]}"
        except Exception as e:
            return "❌ Docker o'rnatilmagan yoki ishlamaydi"
    
    def docker_list_images(self) -> str:
        """List Docker images"""
        try:
            result = subprocess.run(["docker", "images"], capture_output=True, text=True, timeout=10)
            return f"🐳 Docker image lar:\n{result.stdout[:1000]}"
        except Exception as e:
            return "❌ Docker o'rnatilmagan"
    
    def docker_run(self, image: str, command: str = None) -> str:
        """Run a Docker container"""
        try:
            cmd = ["docker", "run", "-d", image]
            if command:
                cmd.extend(["sh", "-c", command])
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
            return f"🐳 Container ishga tushdi:\n{result.stdout[:500]}"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    # ==================== GIT OPERATIONS ====================
    
    def git_status(self, repo_path: str = ".") -> str:
        """Get git status"""
        try:
            result = subprocess.run(["git", "status"], capture_output=True, text=True, 
                                  cwd=repo_path, timeout=10)
            return f"📦 Git Status:\n{result.stdout}"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def git_log(self, repo_path: str = ".", count: int = 10) -> str:
        """Get git log"""
        try:
            result = subprocess.run(["git", "log", f"-{count}", "--oneline"], 
                                  capture_output=True, text=True, cwd=repo_path, timeout=10)
            return f"📦 Git Log:\n{result.stdout}"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def git_add(self, path: str, repo_path: str = ".") -> str:
        """Git add file"""
        try:
            result = subprocess.run(["git", "add", path], capture_output=True, text=True, 
                                  cwd=repo_path, timeout=10)
            return f"✅ Git add: {path}"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def git_commit(self, message: str, repo_path: str = ".") -> str:
        """Git commit"""
        try:
            result = subprocess.run(["git", "commit", "-m", message], 
                                  capture_output=True, text=True, cwd=repo_path, timeout=10)
            return f"✅ Git commit:\n{result.stdout[:500]}"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    # ==================== IMAGE PROCESSING ====================
    
    def image_info(self, filepath: str) -> str:
        """Get image information"""
        try:
            from PIL import Image
            img = Image.open(filepath)
            return f"🖼️ Rasm ma'lumotlari:\n- O'lcham: {img.size}\n- Format: {img.format}\n- Rejim: {img.mode}"
        except ImportError:
            return "❌ Pillow kutubxonasi kerak: pip install pillow"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def resize_image(self, filepath: str, width: int, height: int) -> str:
        """Resize image"""
        try:
            from PIL import Image
            img = Image.open(filepath)
            img = img.resize((width, height))
            new_path = str(Path(filepath).parent / f"resized_{Path(filepath).name}")
            img.save(new_path)
            return f"✅ Rasm o'lchami o'zgartirildi: {new_path}"
        except ImportError:
            return "❌ Pillow kutubxonasi kerak"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    # ==================== WINDOW MANAGEMENT ====================
    
    def list_windows(self) -> str:
        """List open windows"""
        try:
            if os.name == "nt":
                import ctypes
                from ctypes import wintypes
                
                EnumWindows = ctypes.windll.user32.EnumWindows
                EnumWindowsProc = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.POINTER(ctypes.c_int), ctypes.POINTER(ctypes.c_int))
                GetWindowText = ctypes.windll.user32.GetWindowTextW
                GetWindowTextLength = ctypes.windll.user32.GetWindowTextLengthW
                
                windows = []
                
                def callback(hwnd, lParam):
                    length = GetWindowTextLength(hwnd)
                    if length > 0:
                        buff = ctypes.create_unicode_buffer(length + 1)
                        GetWindowText(hwnd, buff, length + 1)
                        windows.append(buff.value)
                    return True
                
                EnumWindows(EnumWindowsProc(callback), 0)
                return f"🪟 Ochiq oynalar:\n" + "\n".join([f"- {w}" for w in windows[:20]])
            else:
                return "❌ Faqat Windows uchun"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    # ==================== COMPRESS/ARCHIVE ====================
    
    def create_zip(self, source: str, destination: str) -> str:
        """Create ZIP archive"""
        try:
            shutil.make_archive(destination, 'zip', source)
            return f"✅ ZIP yaratildi: {destination}.zip"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def extract_zip(self, archive: str, destination: str) -> str:
        """Extract ZIP archive"""
        try:
            shutil.unpack_archive(archive, destination)
            return f"✅ Arxiv ochildi: {destination}"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    # ==================== DOWNLOAD/UPLOAD ====================
    
    def download_file(self, url: str, destination: str) -> str:
        """Download file from URL"""
        try:
            response = requests.get(url, timeout=30, stream=True)
            with open(destination, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
            return f"✅ Yuklab olindi: {destination}"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    # ==================== SYSTEM CONTROL ====================
    
    def restart_system(self) -> str:
        """Restart the computer"""
        try:
            if os.name == "nt":
                subprocess.run(["shutdown", "/r", "/t", "0"], timeout=5)
            else:
                subprocess.run(["reboot"], timeout=5)
            return "♻️ Tizim qayta ishga tushirilmoqda..."
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    def shutdown_system(self) -> str:
        """Shutdown the computer"""
        try:
            if os.name == "nt":
                subprocess.run(["shutdown", "/s", "/t", "0"], timeout=5)
            else:
                subprocess.run(["shutdown", "-h", "now"], timeout=5)
            return "⏻ Tizim o'chirilmoqda..."
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"
    
    # ==================== ENVIRONMENT VARIABLES ====================
    
    def get_env(self, key: str = None) -> str:
        """Get environment variable(s)"""
        if key:
            value = os.environ.get(key, "Topilmadi")
            return f"🔧 {key} = {value}"
        else:
            return f"🔧 Barcha o'zgaruvchilar:\n" + "\n".join([f"{k}={v}" for k, v in list(os.environ.items())[:20]])
    
    def set_env(self, key: str, value: str) -> str:
        """Set environment variable"""
        try:
            os.environ[key] = value
            return f"✅ {key} = {value} o'rnatildi"
        except Exception as e:
            return f"❌ Xatolik: {str(e)}"


# Tool execution dispatcher - NO LIMITS VERSION
def execute_tool(tool_name: str, params: Dict[str, Any], tools_engine: ToolsEngine) -> str:
    """Execute a specific tool based on name and parameters"""
    
    tool_map = {
        "read_file": tools_engine.read_file,
        "write_file": tools_engine.write_file,
        "delete_file": tools_engine.delete_file,
        "list_directory": tools_engine.list_directory,
        "search_files": tools_engine.search_files,
        "execute_command": tools_engine.execute_command,
        "web_search": tools_engine.web_search,
        "scrape_website": tools_engine.scrape_website,
        "execute_code": tools_engine.execute_code,
        "analyze_data": tools_engine.analyze_data,
        "launch_application": tools_engine.launch_application,
        "get_system_info": tools_engine.get_system_info,
        "get_current_time": tools_engine.get_current_time,
        "check_password_strength": tools_engine.check_password_strength,
        "ping_host": tools_engine.ping_host,
    }
    
    tool_func = tool_map.get(tool_name)
    if tool_func:
        return tool_func(**params)
    else:
        return f"❌ Noma'lum tool: {tool_name}"